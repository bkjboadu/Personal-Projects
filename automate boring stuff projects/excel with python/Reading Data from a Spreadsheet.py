import pprint,openpyxl

from openpyxl.utils import get_column_letter,column_index_from_string
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.active
alldata = {}
for row in range(2,sheet.max_row+1):
    state = sheet.cell(row=row,column=2).value
    county = sheet.cell(row=row,column=3).value
    pop = sheet.cell(row=row,column=4).value
    alldata.setdefault(state,{})
    alldata[state].setdefault(county,{'tract':0,'pop':0})
    alldata[state][county]['tract'] += 1
    alldata[state][county]['pop'] += pop

print('Writing results....')
resultFile = open('censuspop.py','w')
resultFile.write('alldata = ' + pprint.pformat(alldata))
resultFile.close()
print('Done')
