import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
new_prices = {'Celery':1.19,'Garlic':3.07,'Lemon':1.27}
for i in range(1,sheet.max_row + 1):
    product = sheet.cell(column=1,row=i).value
    if product in new_prices.keys():
        sheet.cell(column=2,row=i).value = new_prices[product]

wb.save('produce_done.xlsx')